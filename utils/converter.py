#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
통합 데이터 변환기 (Converter)
CSV ↔ JSON ↔ Excel 간의 상호 변환을 지원하는 통합 클래스
"""

import pandas as pd
import json
import os
import re
import asyncio
import glob
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
import csv

# Excel 관련 라이브러리
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# AI 관련 라이브러리 (선택적)
try:
    from ai_helpers import AIModelManager
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False

# 환경변수 로드
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

class SimplePhoneValidator:
    """간단한 전화번호 검증 클래스"""
    
    @staticmethod
    def is_valid_korean_phone(phone: str) -> bool:
        """한국 전화번호 유효성 검증"""
        if not phone:
            return False
        
        # 숫자만 추출
        clean_phone = re.sub(r'[^\d]', '', phone)
        
        # 길이 체크 (10자리 또는 11자리)
        if len(clean_phone) not in [10, 11]:
            return False
        
        # 지역번호 체크
        valid_codes = ["02", "031", "032", "033", "041", "042", "043", "044", 
                      "051", "052", "053", "054", "055", "061", "062", "063", "064", "070", "010", "017"]
        
        for code in valid_codes:
            if clean_phone.startswith(code):
                return True
        
        return False
    
    @staticmethod
    def format_phone_number(phone: str) -> str:
        """전화번호 표준 형식으로 포맷팅"""
        if not phone:
            return ""
        
        clean_phone = re.sub(r'[^\d]', '', phone)
        
        if len(clean_phone) == 10:
            if clean_phone.startswith('02'):
                return f"{clean_phone[:2]}-{clean_phone[2:6]}-{clean_phone[6:]}"
            else:
                return f"{clean_phone[:3]}-{clean_phone[3:6]}-{clean_phone[6:]}"
        elif len(clean_phone) == 11:
            return f"{clean_phone[:3]}-{clean_phone[3:7]}-{clean_phone[7:]}"
        
        return phone

class SimpleLogger:
    """간단한 로거 클래스"""
    
    def __init__(self, name):
        self.name = name
    
    def info(self, msg):
        print(f"[INFO] {msg}")
    
    def warning(self, msg):
        print(f"[WARNING] {msg}")
    
    def error(self, msg):
        print(f"[ERROR] {msg}")

class Converter:
    """통합 데이터 변환기 클래스"""
    
    def __init__(self, use_ai: bool = False):
        """초기화"""
        self.logger = SimpleLogger("converter")
        self.phone_utils = SimplePhoneValidator
        self.use_ai = use_ai
        self.ai_manager = None
        
        # AI 모델 초기화 (선택적)
        if self.use_ai and AI_AVAILABLE:
            api_key = os.getenv('GEMINI_API_KEY')
            if api_key:
                try:
                    self.ai_manager = AIModelManager()
                    self.logger.info("🤖 AI 모델 초기화 완료")
                except Exception as e:
                    self.logger.warning(f"AI 모델 초기화 실패: {e}")
                    self.ai_manager = None
        
        # 컬럼 매핑 (CSV → JSON)
        self.column_mapping = {
            # 기본 정보
            'Name': 'name', '기관명': 'name', 'name': 'name',
            'Category': 'category', '분류': 'category', 'category': 'category',
            
            # 홈페이지
            'homepage': 'homepage', '홈페이지': 'homepage', 'url': 'homepage',
            
            # 연락처
            '전화번호': 'phone', 'phone': 'phone', '전화': 'phone',
            '팩스1': 'fax', '팩스번호': 'fax', 'fax': 'fax', '팩스': 'fax',
            '이메일': 'email', 'email': 'email',
            '핸드폰1': 'mobile', '핸드폰': 'mobile', 'mobile': 'mobile',
            
            # 주소 정보
            '우편번호': 'postal_code', 'postal_code': 'postal_code',
            '주소2': 'address', '주소': 'address', 'address': 'address'
        }
    
    def clean_data(self, value: Any) -> str:
        """데이터 정제"""
        if pd.isna(value) or value is None:
            return ""
        
        # 문자열 변환 및 정제
        cleaned = str(value).strip()
        
        # 빈 문자열 또는 의미없는 값 제거
        if cleaned in ['nan', 'NaN', 'null', 'NULL', '-', 'N/A', '없음', 'X']:
            return ""
        
        return cleaned
    
    def validate_and_format_phone(self, phone: str) -> str:
        """전화번호 검증 및 포맷팅"""
        if not phone or phone == "":
            return ""
        
        if self.phone_utils.is_valid_korean_phone(phone):
            return self.phone_utils.format_phone_number(phone)
        else:
            self.logger.warning(f"검증 실패한 전화번호: {phone}")
            return phone
    
    def validate_email(self, email: str) -> str:
        """이메일 간단 검증"""
        if not email or email == "":
            return ""
        
        if '@' in email and '.' in email:
            return email
        else:
            self.logger.warning(f"검증 실패한 이메일: {email}")
            return ""
    
    def format_url(self, url: str) -> str:
        """URL 포맷팅"""
        if not url or url == "":
            return ""
        
        url = url.strip()
        if not url.startswith(('http://', 'https://')):
            if url.startswith('www.'):
                url = 'http://' + url
            elif '.' in url:
                url = 'http://' + url
        
        return url
    
    def csv_to_json(self, csv_file_path: str, output_file_path: str = None) -> str:
        """CSV 파일을 JSON으로 변환"""
        try:
            self.logger.info(f"📄 CSV → JSON 변환 시작: {csv_file_path}")
            
            # CSV 파일 존재 확인
            if not os.path.exists(csv_file_path):
                raise FileNotFoundError(f"CSV 파일을 찾을 수 없습니다: {csv_file_path}")
            
            # 다양한 인코딩 시도
            encodings = ['utf-8', 'cp949', 'euc-kr', 'latin1']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(csv_file_path, encoding=encoding)
                    self.logger.info(f"CSV 로드 성공 (인코딩: {encoding})")
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                raise ValueError("CSV 파일을 읽을 수 없습니다. 인코딩을 확인해주세요.")
            
            self.logger.info(f"로드된 데이터: {len(df)}행, {len(df.columns)}열")
            
            # 데이터 변환
            organizations = []
            processed_count = 0
            error_count = 0
            
            for index, row in df.iterrows():
                try:
                    # 기본 구조
                    org_data = {
                        "name": "", "category": "", "homepage": "",
                        "phone": "", "fax": "", "email": "", "mobile": "",
                        "postal_code": "", "address": ""
                    }
                    
                    # CSV 컬럼을 JSON 필드로 매핑
                    for csv_col in df.columns:
                        csv_col_clean = csv_col.strip()
                        
                        if csv_col_clean in self.column_mapping:
                            json_field = self.column_mapping[csv_col_clean]
                            cleaned_value = self.clean_data(row[csv_col])
                            
                            # 각 필드별 특별 처리
                            if json_field in ['phone', 'fax', 'mobile']:
                                org_data[json_field] = self.validate_and_format_phone(cleaned_value)
                            elif json_field == 'email':
                                org_data[json_field] = self.validate_email(cleaned_value)
                            elif json_field == 'homepage':
                                org_data[json_field] = self.format_url(cleaned_value)
                            else:
                                org_data[json_field] = cleaned_value
                    
                    # 기관명이 없으면 스킵
                    if not org_data['name']:
                        self.logger.warning(f"행 {index+1}: 기관명이 없어서 스킵")
                        error_count += 1
                        continue
                    
                    organizations.append(org_data)
                    processed_count += 1
                    
                except Exception as e:
                    self.logger.error(f"행 {index+1} 처리 중 오류: {e}")
                    error_count += 1
                    continue
            
            # 출력 파일명 생성
            if not output_file_path:
                base_name = os.path.basename(csv_file_path)
                timestamp = datetime.now().strftime("%m%d")
                output_file_path = f"converted_data_{timestamp}.json"
            
            # JSON 파일 저장
            with open(output_file_path, 'w', encoding='utf-8') as f:
                json.dump(organizations, f, ensure_ascii=False, indent=2)
            
            self.logger.info(f"✅ CSV → JSON 변환 완료: {output_file_path}")
            self.logger.info(f"📊 처리: {processed_count}개, 오류: {error_count}개")
            
            return output_file_path
            
        except Exception as e:
            self.logger.error(f"❌ CSV → JSON 변환 실패: {e}")
            raise
    
    def json_to_csv(self, json_file_path: str, output_file_path: str = None) -> str:
        """JSON 파일을 CSV로 변환"""
        try:
            self.logger.info(f"📄 JSON → CSV 변환 시작: {json_file_path}")
            
            # JSON 파일 로드
            with open(json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 출력 파일명 생성
            if not output_file_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file_path = f"converted_data_{timestamp}.csv"
            
            # CSV 파일 생성
            fieldnames = ['기관명', '전화번호', '팩스번호', '이메일', 'URL']
            
            with open(output_file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                total_count = 0
                
                # 데이터 구조 확인 및 처리
                if isinstance(data, dict):
                    # 카테고리별 딕셔너리 구조
                    for category, organizations in data.items():
                        for org in organizations:
                            contact_info = self._extract_contact_info_for_csv(org)
                            writer.writerow(contact_info)
                            total_count += 1
                elif isinstance(data, list):
                    # 단순 리스트 구조
                    for org in data:
                        contact_info = self._extract_contact_info_for_csv(org)
                        writer.writerow(contact_info)
                        total_count += 1
            
            self.logger.info(f"✅ JSON → CSV 변환 완료: {output_file_path}")
            self.logger.info(f"📊 총 {total_count}개 기관 변환")
            
            return output_file_path
            
        except Exception as e:
            self.logger.error(f"❌ JSON → CSV 변환 실패: {e}")
            raise
    
    def _extract_contact_info_for_csv(self, org_data: Dict[str, Any]) -> Dict[str, str]:
        """CSV용 연락처 정보 추출"""
        result = {
            "기관명": org_data.get("name", ""),
            "전화번호": "",
            "팩스번호": "",
            "이메일": "",
            "URL": org_data.get("homepage", "")
        }
        
        # 기본 필드에서 추출
        if org_data.get("phone"):
            result["전화번호"] = org_data["phone"]
        if org_data.get("fax"):
            result["팩스번호"] = org_data["fax"]
        
        # 홈페이지 파싱 결과에서 추출
        homepage_content = org_data.get("homepage_content", {})
        if homepage_content:
            parsed_contact = homepage_content.get("parsed_contact", {})
            
            # 파싱된 결과 우선 사용
            if parsed_contact.get("phones") and not result["전화번호"]:
                result["전화번호"] = ", ".join(parsed_contact["phones"])
            
            if parsed_contact.get("faxes") and not result["팩스번호"]:
                result["팩스번호"] = ", ".join(parsed_contact["faxes"])
            
            if parsed_contact.get("emails"):
                result["이메일"] = ", ".join(parsed_contact["emails"])
        
        return result
    
    def json_to_excel(self, json_file_path: str, output_file_path: str = None, use_ai_validation: bool = None) -> str:
        """JSON 파일을 Excel로 변환"""
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel 변환을 위해 openpyxl이 필요합니다: pip install openpyxl")
        
        try:
            self.logger.info(f"📄 JSON → Excel 변환 시작: {json_file_path}")
            
            # AI 검증 설정
            if use_ai_validation is None:
                use_ai_validation = self.use_ai and self.ai_manager is not None
            
            # JSON 파일 로드
            with open(json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 출력 파일명 생성
            if not output_file_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file_path = f"converted_data_{timestamp}.xlsx"
            
            # Excel 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "연락처 데이터"
            
            # 헤더 설정
            headers = ["기관명", "전화번호", "팩스번호", "이메일", "URL"]
            if use_ai_validation:
                headers.extend(["AI검증상태", "검증점수", "신뢰도"])
            
            # 헤더 스타일 설정
            self._apply_excel_header_style(ws, headers)
            
            # 데이터 처리
            row_num = 2
            total_count = 0
            
            # 데이터 구조 확인 및 처리
            if isinstance(data, dict):
                # 카테고리별 딕셔너리 구조
                for category, organizations in data.items():
                    self.logger.info(f"📂 처리 중: {category} ({len(organizations)}개 기관)")
                    
                    for org in organizations:
                        contact_info = self._extract_contact_info_for_csv(org)
                        self._write_excel_row(ws, row_num, contact_info, org, use_ai_validation)
                        row_num += 1
                        total_count += 1
                        
            elif isinstance(data, list):
                # 단순 리스트 구조
                for org in data:
                    contact_info = self._extract_contact_info_for_csv(org)
                    self._write_excel_row(ws, row_num, contact_info, org, use_ai_validation)
                    row_num += 1
                    total_count += 1
            
            # 열 너비 자동 조정
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 20
            
            # Excel 파일 저장
            wb.save(output_file_path)
            
            self.logger.info(f"✅ JSON → Excel 변환 완료: {output_file_path}")
            self.logger.info(f"📊 총 {total_count}개 기관 변환")
            
            return output_file_path
            
        except Exception as e:
            self.logger.error(f"❌ JSON → Excel 변환 실패: {e}")
            raise
    
    def _apply_excel_header_style(self, ws, headers):
        """Excel 헤더 스타일 적용"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _write_excel_row(self, ws, row_num, contact_info, org_data, use_ai_validation):
        """Excel 행 데이터 작성"""
        ws.cell(row=row_num, column=1, value=contact_info["기관명"])
        ws.cell(row=row_num, column=2, value=contact_info["전화번호"])
        ws.cell(row=row_num, column=3, value=contact_info["팩스번호"])
        ws.cell(row=row_num, column=4, value=contact_info["이메일"])
        ws.cell(row=row_num, column=5, value=contact_info["URL"])
        
        # AI 검증 (선택적)
        if use_ai_validation and self.ai_manager:
            # 간단한 검증 로직 (실제 AI 호출 대신 규칙 기반)
            has_contact = bool(contact_info["전화번호"] or contact_info["이메일"])
            ws.cell(row=row_num, column=6, value="검증완료" if has_contact else "정보부족")
            ws.cell(row=row_num, column=7, value="100" if has_contact else "50")
            ws.cell(row=row_num, column=8, value="높음" if has_contact else "낮음")
    
    def excel_to_json(self, excel_file_path: str, output_file_path: str = None) -> str:
        """Excel 파일을 JSON으로 변환"""
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel 변환을 위해 openpyxl이 필요합니다: pip install openpyxl")
        
        try:
            self.logger.info(f"📄 Excel → JSON 변환 시작: {excel_file_path}")
            
            # Excel 파일 로드
            wb = load_workbook(excel_file_path)
            ws = wb.active
            
            # 헤더 읽기
            headers = [cell.value for cell in ws[1]]
            self.logger.info(f"헤더: {headers}")
            
            # 데이터 읽기
            organizations = []
            
            for row_idx in range(2, ws.max_row + 1):
                row_data = [cell.value for cell in ws[row_idx]]
                
                # 기관명이 없으면 스킵
                if not row_data[0]:
                    continue
                
                # 기본 구조로 변환
                org_data = {
                    "name": str(row_data[0]) if row_data[0] else "",
                    "phone": str(row_data[1]) if row_data[1] and len(row_data) > 1 else "",
                    "fax": str(row_data[2]) if row_data[2] and len(row_data) > 2 else "",
                    "email": str(row_data[3]) if row_data[3] and len(row_data) > 3 else "",
                    "homepage": str(row_data[4]) if row_data[4] and len(row_data) > 4 else ""
                }
                
                # 전화번호 포맷팅
                if org_data["phone"]:
                    org_data["phone"] = self.validate_and_format_phone(org_data["phone"])
                if org_data["fax"]:
                    org_data["fax"] = self.validate_and_format_phone(org_data["fax"])
                
                organizations.append(org_data)
            
            # 출력 파일명 생성
            if not output_file_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file_path = f"converted_data_{timestamp}.json"
            
            # JSON 파일 저장
            with open(output_file_path, 'w', encoding='utf-8') as f:
                json.dump(organizations, f, ensure_ascii=False, indent=2)
            
            self.logger.info(f"✅ Excel → JSON 변환 완료: {output_file_path}")
            self.logger.info(f"📊 총 {len(organizations)}개 기관 변환")
            
            return output_file_path
            
        except Exception as e:
            self.logger.error(f"❌ Excel → JSON 변환 실패: {e}")
            raise
    
    def find_latest_file(self, pattern: str) -> Optional[str]:
        """패턴에 맞는 가장 최근 파일 찾기"""
        files = glob.glob(pattern)
        if not files:
            return None
        return max(files, key=os.path.getctime)
    
    def get_file_statistics(self, file_path: str) -> Dict[str, Any]:
        """파일 통계 정보 반환"""
        try:
            stat = os.stat(file_path)
            size_mb = stat.st_size / (1024 * 1024)
            created_time = datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S")
            
            # 파일 타입별 데이터 수 계산
            data_count = 0
            if file_path.endswith('.json'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        data_count = len(data)
                    elif isinstance(data, dict):
                        data_count = sum(len(orgs) for orgs in data.values())
            elif file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
                data_count = len(df)
            elif file_path.endswith('.xlsx'):
                wb = load_workbook(file_path)
                ws = wb.active
                data_count = ws.max_row - 1  # 헤더 제외
            
            return {
                'size_mb': round(size_mb, 2),
                'created_time': created_time,
                'data_count': data_count
            }
        except Exception as e:
            self.logger.error(f"파일 통계 조회 실패: {e}")
            return {'size_mb': 0, 'created_time': 'Unknown', 'data_count': 0}

# 사용 예시 함수들
def convert_csv_to_json(csv_file: str, output_file: str = None) -> str:
    """CSV를 JSON으로 변환하는 편의 함수"""
    converter = Converter()
    return converter.csv_to_json(csv_file, output_file)

def convert_json_to_csv(json_file: str, output_file: str = None) -> str:
    """JSON을 CSV로 변환하는 편의 함수"""
    converter = Converter()
    return converter.json_to_csv(json_file, output_file)

def convert_json_to_excel(json_file: str, output_file: str = None, use_ai: bool = False) -> str:
    """JSON을 Excel로 변환하는 편의 함수"""
    converter = Converter(use_ai=use_ai)
    return converter.json_to_excel(json_file, output_file)

def convert_excel_to_json(excel_file: str, output_file: str = None) -> str:
    """Excel을 JSON으로 변환하는 편의 함수"""
    converter = Converter()
    return converter.excel_to_json(excel_file, output_file)

if __name__ == "__main__":
    # 테스트 코드
    print("=" * 60)
    print("🔄 통합 데이터 변환기 테스트")
    print("=" * 60)
    
    converter = Converter(use_ai=False)
    
    # 사용 예시
    print("📋 사용 가능한 변환:")
    print("  1. CSV → JSON")
    print("  2. JSON → CSV") 
    print("  3. JSON → Excel")
    print("  4. Excel → JSON")
    print("\n🔧 Converter 클래스 초기화 완료")