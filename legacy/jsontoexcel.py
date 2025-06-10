#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JSON to Excel 변환기 (Gemini AI 검증 포함)
홈페이지 크롤링 결과 JSON 파일을 Excel 파일로 변환하는 스크립트
"""

import json
import os
import sys
import glob
import re
import asyncio
from datetime import datetime
from typing import Dict, List, Any, Optional
import traceback

# 상위 디렉토리를 Python path에 추가 (ai_helpers 모듈 접근용)
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

# Excel 관련 import
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 환경변수 로드
try:
    from dotenv import load_dotenv
    load_dotenv()
    print("✅ .env 파일 로드 완료")
except ImportError:
    print("⚠️ python-dotenv 모듈이 없습니다. 환경변수를 직접 설정해주세요.")

# AI 도우미 모듈 import (선택적)
try:
    from ai_helpers import AIManager
    print("✅ ai_helpers 모듈 로드 완료")
    AI_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ ai_helpers 모듈이 없어 AI 검증 기능 비활성화: {e}")
    AIManager = None
    AI_AVAILABLE = False

class ContactDataExtractor:
    """연락처 데이터 추출 및 검증 클래스"""
    
    def __init__(self):
        """초기화"""
        self.ai_manager = None
        
        # .env에서 API 키 확인
        api_key = os.getenv('GEMINI_API_KEY')
        
        if AI_AVAILABLE and api_key:
            try:
                # AIModelManager에 API 키 전달 (필요시)
                self.ai_manager = AIManager()
                print("🤖 AI 모델 초기화 완료 (.env에서 API 키 로드)")
            except Exception as e:
                print(f"⚠️ AI 모델 초기화 실패: {e}")
                self.ai_manager = None
        elif AI_AVAILABLE and not api_key:
            print("⚠️ .env에서 GEMINI_API_KEY를 찾을 수 없습니다. AI 검증 기능 비활성화")
        else:
            print("⚠️ ai_helpers 모듈이 없어 AI 검증 기능 비활성화")
    
    def find_latest_enhanced_file() -> str:
        """가장 최근의 churches_enhanced_final_*.json 파일 찾기"""
        try:
            # churches_enhanced_final_*.json 패턴으로 파일 찾기
            pattern = "churches_enhanced_final_*.json"
            files = glob.glob(pattern)
            
            if not files:
                print(f"❌ {pattern} 패턴의 파일을 찾을 수 없습니다.")
                return ""
            
            # 가장 최근 파일 선택 (파일 생성 시간 기준)
            latest_file = max(files, key=os.path.getctime)
            
            print(f"🔍 발견된 파일들:")
            for file in sorted(files, key=os.path.getctime, reverse=True):
                marker = " ← 선택됨" if file == latest_file else ""
                creation_time = datetime.fromtimestamp(os.path.getctime(file)).strftime("%Y-%m-%d %H:%M:%S")
                print(f"  📄 {file} (생성: {creation_time}){marker}")
            
            return latest_file
            
        except Exception as e:
            print(f"❌ 파일 검색 중 오류: {e}")
            return ""

    def find_latest_json_file() -> str:
        """가장 최근의 JSON 파일 찾기 (수정된 버전)"""
        try:
            # 현재 디렉토리에서 패턴 검색
            patterns = [
                "churches_enhanced_final_*.json",  # advcrawler.py 결과
                "raw_data_with_homepages_*.json",  # url_extractor 결과  
                "undefined_converted_*.json"       # 원본 데이터
            ]
            
            all_files = []
            for pattern in patterns:
                files = glob.glob(pattern)
                all_files.extend(files)
            
            if not all_files:
                print("❌ JSON 파일을 찾을 수 없습니다.")
                return ""
            
            # 파일 수정 시간 기준으로 최신 파일 선택
            latest_file = max(all_files, key=os.path.getctime)
            print(f"🔍 찾은 최신 파일: {latest_file}")
            
            return latest_file
            
        except Exception as e:
            print(f"❌ 파일 검색 중 오류: {e}")
            return ""
        
    def filter_news_content(self, text: str) -> bool:
        """news 키워드가 포함된 내용 필터링"""
        if not text:
            return False
        
        # news 관련 키워드 체크 (대소문자 무관)
        news_keywords = ['news', 'newsletter', 'newsroom', 'press', 'media']
        text_lower = text.lower()
        
        for keyword in news_keywords:
            if keyword in text_lower:
                return True
        return False
    
    def extract_contact_info(self, org_data: Dict[str, Any]) -> Dict[str, str]:
        """기관 데이터에서 연락처 정보 추출"""
        result = {
            "기관명": org_data.get("name", ""),
            "전화번호": "",
            "fax번호": "",
            "이메일": "",
            "url": org_data.get("homepage", "")
        }
        
        # 기본 필드에서 추출
        if org_data.get("phone"):
            result["전화번호"] = org_data["phone"]
        if org_data.get("fax"):
            result["fax번호"] = org_data["fax"]
        
        # 홈페이지 파싱 결과에서 추출
        homepage_content = org_data.get("homepage_content", {})
        if homepage_content:
            parsed_contact = homepage_content.get("parsed_contact", {})
            
            # 전화번호 추출 (파싱된 결과 우선)
            if parsed_contact.get("phones") and not result["전화번호"]:
                phones = [phone for phone in parsed_contact["phones"] if not self.filter_news_content(phone)]
                if phones:
                    result["전화번호"] = ", ".join(phones)
            
            # 팩스번호 추출 (파싱된 결과 우선)
            if parsed_contact.get("faxes") and not result["fax번호"]:
                faxes = [fax for fax in parsed_contact["faxes"] if not self.filter_news_content(fax)]
                if faxes:
                    result["fax번호"] = ", ".join(faxes)
            
            # 이메일 추출 (news 필터링)
            if parsed_contact.get("emails"):
                emails = [email for email in parsed_contact["emails"] if not self.filter_news_content(email)]
                if emails:
                    result["이메일"] = ", ".join(emails)
        
        # URL news 필터링
        if result["url"] and self.filter_news_content(result["url"]):
            result["url"] = ""
        
        return result
    
    def validate_contact_consistency(self, org_data: Dict[str, Any], extracted_data: Dict[str, str]) -> Dict[str, Any]:
        """contact_info와 parsed_contact 데이터 일치성 검증"""
        validation_result = {
            "is_consistent": True,
            "issues": [],
            "confidence_score": 1.0
        }
        
        homepage_content = org_data.get("homepage_content", {})
        contact_info = homepage_content.get("contact_info", "")
        parsed_contact = homepage_content.get("parsed_contact", {})
        
        if not contact_info and not parsed_contact:
            validation_result["issues"].append("연락처 정보 없음")
            validation_result["confidence_score"] = 0.0
            return validation_result
        
        # 기본 일치성 검증
        issues = []
        
        # 전화번호 검증
        if extracted_data["전화번호"]:
            if contact_info and extracted_data["전화번호"] not in contact_info:
                issues.append(f"전화번호 불일치: {extracted_data['전화번호']}")
        
        # 팩스번호 검증
        if extracted_data["fax번호"]:
            if contact_info and extracted_data["fax번호"] not in contact_info:
                issues.append(f"팩스번호 불일치: {extracted_data['fax번호']}")
        
        # 이메일 검증
        if extracted_data["이메일"]:
            if contact_info and extracted_data["이메일"] not in contact_info:
                issues.append(f"이메일 불일치: {extracted_data['이메일']}")
        
        if issues:
            validation_result["is_consistent"] = False
            validation_result["issues"] = issues
            validation_result["confidence_score"] = max(0.1, 1.0 - len(issues) * 0.3)
        
        return validation_result
    
    async def ai_validate_contact_data(self, org_data: Dict[str, Any], extracted_data: Dict[str, str]) -> Dict[str, Any]:
        """Gemini AI를 사용한 연락처 데이터 검증"""
        if not self.ai_manager:
            return {"ai_validation": "AI 검증 불가능"}
        
        try:
            homepage_content = org_data.get("homepage_content", {})
            contact_info = homepage_content.get("contact_info", "")
            parsed_contact = homepage_content.get("parsed_contact", {})
            
            # AI 검증용 프롬프트 구성
            prompt_template = """
연락처 정보 검증 전문가로서, 다음 데이터의 일치성을 검증해주세요.

**기관명**: {org_name}

**원본 연락처 정보 (contact_info)**:
{contact_info}

**파싱된 연락처 정보 (parsed_contact)**:
- 전화번호: {phones}
- 팩스번호: {faxes}
- 이메일: {emails}

**추출된 최종 데이터**:
- 전화번호: {final_phone}
- 팩스번호: {final_fax}
- 이메일: {final_email}

다음 사항을 검증해주세요:
1. 원본 정보와 파싱된 정보의 일치성
2. 추출된 최종 데이터의 정확성
3. 누락되거나 잘못된 정보가 있는지
4. news 관련 이메일/연락처가 제대로 필터링되었는지

검증 결과를 다음 형식으로 답변해주세요:
- **일치성 점수**: (0-100점)
- **주요 문제점**: (있다면 나열)
- **권장 수정사항**: (있다면 제안)
- **신뢰도**: (높음/보통/낮음)

{text_content}
"""
            
            # 프롬프트 데이터 구성
            prompt_data = prompt_template.format(
                org_name=extracted_data["기관명"],
                contact_info=contact_info or "정보 없음",
                phones=", ".join(parsed_contact.get("phones", [])) or "없음",
                faxes=", ".join(parsed_contact.get("faxes", [])) or "없음", 
                emails=", ".join(parsed_contact.get("emails", [])) or "없음",
                final_phone=extracted_data["전화번호"] or "없음",
                final_fax=extracted_data["fax번호"] or "없음",
                final_email=extracted_data["이메일"] or "없음",
                text_content=""
            )
            
            # AI 검증 실행
            ai_response = await self.ai_manager.extract_with_gemini(
                text_content="",  # 이미 프롬프트에 포함됨
                prompt_template=prompt_data
            )
            
            # 마크다운 응답 파싱
            parsed_response = self.parse_markdown_response(ai_response)
            
            return {
                "ai_validation": "완료",
                "ai_response": ai_response,
                "parsed_validation": parsed_response
            }
            
        except Exception as e:
            print(f"⚠️ AI 검증 중 오류: {e}")
            return {"ai_validation": f"오류: {str(e)}"}
    
    def parse_markdown_response(self, markdown_text: str) -> Dict[str, str]:
        """마크다운 형식의 AI 응답 파싱"""
        if not markdown_text:
            return {}
        
        parsed = {}
        
        # 일치성 점수 추출
        score_match = re.search(r'일치성 점수.*?(\d+)', markdown_text)
        if score_match:
            parsed["consistency_score"] = score_match.group(1)
        
        # 주요 문제점 추출
        problems_match = re.search(r'주요 문제점.*?:(.*?)(?=\*\*|$)', markdown_text, re.DOTALL)
        if problems_match:
            parsed["problems"] = problems_match.group(1).strip()
        
        # 권장 수정사항 추출
        recommendations_match = re.search(r'권장 수정사항.*?:(.*?)(?=\*\*|$)', markdown_text, re.DOTALL)
        if recommendations_match:
            parsed["recommendations"] = recommendations_match.group(1).strip()
        
        # 신뢰도 추출
        reliability_match = re.search(r'신뢰도.*?:(.*?)(?=\*\*|$)', markdown_text)
        if reliability_match:
            parsed["reliability"] = reliability_match.group(1).strip()
        
        return parsed
    
    def chunk_json_data(self, data: List[Dict], chunk_size: int = 50) -> List[List[Dict]]:
        """JSON 데이터를 청크 단위로 분할 (context limit 대응)"""
        chunks = []
        
        # 리스트를 chunk_size 단위로 분할
        for i in range(0, len(data), chunk_size):
            chunk = data[i:i + chunk_size]
            chunks.append(chunk)
        
        print(f"📦 데이터를 {len(chunks)}개 청크로 분할 (청크당 최대 {chunk_size}개 기관)")
        return chunks
    
    async def process_json_to_excel(self, json_file_path: str, excel_file_path: str, use_ai_validation: bool = True) -> bool:
        """JSON 파일을 Excel 파일로 변환 (AI 검증 포함)"""
        print(f"🔄 JSON to Excel 변환 시작...")
        print(f"📂 입력 파일: {json_file_path}")
        print(f"💾 출력 파일: {excel_file_path}")
        print(f"🤖 AI 검증: {'활성화' if use_ai_validation and self.ai_manager else '비활성화'}")
        
        try:
            # JSON 파일 로드
            with open(json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            print(f"✅ JSON 파일 로드 완료: {len(data)}개 기관")
            
            # Excel 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "연락처 데이터"
            
            # 헤더 설정
            headers = ["기관명", "전화번호", "fax번호", "이메일", "url"]
            if use_ai_validation and self.ai_manager:
                headers.extend(["일치성검증", "AI검증점수", "신뢰도", "문제점"])
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 데이터 처리
            row_num = 2
            total_count = 0
            ai_validation_count = 0
            
            # 청크 단위로 처리 (메모리 효율성)
            chunks = self.chunk_json_data(data, chunk_size=50)
            
            for chunk_idx, chunk in enumerate(chunks):
                print(f"📦 청크 {chunk_idx + 1}/{len(chunks)} 처리 중...")
                
                for org in chunk:
                    # 기본 연락처 정보 추출
                    contact_info = self.extract_contact_info(org)
                    
                    # 기본 일치성 검증
                    validation_result = self.validate_contact_consistency(org, contact_info)
                    
                    # Excel 행에 기본 데이터 입력
                    ws.cell(row=row_num, column=1, value=contact_info["기관명"])
                    ws.cell(row=row_num, column=2, value=contact_info["전화번호"])
                    ws.cell(row=row_num, column=3, value=contact_info["fax번호"])
                    ws.cell(row=row_num, column=4, value=contact_info["이메일"])
                    ws.cell(row=row_num, column=5, value=contact_info["url"])
                    
                    # AI 검증 (선택적)
                    if use_ai_validation and self.ai_manager and not validation_result["is_consistent"]:
                        try:
                            ai_result = await self.ai_validate_contact_data(org, contact_info)
                            parsed_ai = ai_result.get("parsed_validation", {})
                            
                            ws.cell(row=row_num, column=6, value="AI검증완료" if ai_result.get("ai_validation") == "완료" else "검증실패")
                            ws.cell(row=row_num, column=7, value=parsed_ai.get("consistency_score", ""))
                            ws.cell(row=row_num, column=8, value=parsed_ai.get("reliability", ""))
                            ws.cell(row=row_num, column=9, value=parsed_ai.get("problems", ""))
                            
                            ai_validation_count += 1
                            
                        except Exception as ai_err:
                            print(f"⚠️ AI 검증 오류 ({contact_info['기관명']}): {ai_err}")
                            ws.cell(row=row_num, column=6, value="AI검증오류")
                    
                    elif use_ai_validation and self.ai_manager:
                        ws.cell(row=row_num, column=6, value="검증통과")
                        ws.cell(row=row_num, column=7, value="100")
                        ws.cell(row=row_num, column=8, value="높음")
                    
                    row_num += 1
                    total_count += 1
                    
                    # 진행 상황 표시 (50개마다)
                    if total_count % 50 == 0:
                        print(f"   📝 {total_count}개 기관 처리 완료...")
            
            # 열 너비 자동 조정
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 20
            
            # Excel 파일 저장
            wb.save(excel_file_path)
            
            print(f"🎉 Excel 변환 완료!")
            print(f"📊 총 {total_count}개 기관 데이터 변환됨")
            if use_ai_validation and self.ai_manager:
                print(f"🤖 AI 검증 수행: {ai_validation_count}개 기관")
            print(f"💾 저장 위치: {excel_file_path}")
            
            return True
            
        except Exception as e:
            print(f"❌ 변환 중 오류 발생: {e}")
            print(f"🔍 상세 오류: {traceback.format_exc()}")
            return False
    
    def preview_excel_data(self, excel_file_path: str, num_rows: int = 5):
        """Excel 파일 미리보기"""
        print(f"\n📋 Excel 파일 미리보기 (상위 {num_rows}개 행):")
        print("=" * 100)
        
        try:
            wb = load_workbook(excel_file_path)
            ws = wb.active
            
            # 헤더 출력
            headers = [cell.value for cell in ws[1]]
            header_line = " | ".join([f"{str(header)[:15]:15}" for header in headers])
            print(header_line)
            print("-" * len(header_line))
            
            # 데이터 행 출력
            for row_idx in range(2, min(num_rows + 2, ws.max_row + 1)):
                row_data = [cell.value for cell in ws[row_idx]]
                row_line = " | ".join([f"{str(data)[:15] if data else '':15}" for data in row_data])
                print(row_line)
            
            print("=" * 100)
            
        except Exception as e:
            print(f"❌ 미리보기 중 오류: {e}")
    
    def count_excel_statistics(self, excel_file_path: str):
        """Excel 데이터 통계 출력"""
        print(f"\n📊 데이터 통계:")
        print("=" * 50)
        
        try:
            wb = load_workbook(excel_file_path)
            ws = wb.active
            
            total_count = ws.max_row - 1  # 헤더 제외
            phone_count = 0
            email_count = 0
            fax_count = 0
            url_count = 0
            
            for row_idx in range(2, ws.max_row + 1):
                phone = ws.cell(row=row_idx, column=2).value
                fax = ws.cell(row=row_idx, column=3).value
                email = ws.cell(row=row_idx, column=4).value
                url = ws.cell(row=row_idx, column=5).value
                
                if phone and str(phone).strip():
                    phone_count += 1
                if fax and str(fax).strip():
                    fax_count += 1
                if email and str(email).strip():
                    email_count += 1
                if url and str(url).strip():
                    url_count += 1
            
            print(f"📈 총 기관 수: {total_count}")
            print(f"📞 전화번호 보유: {phone_count}개 ({phone_count/total_count*100:.1f}%)")
            print(f"📠 팩스번호 보유: {fax_count}개 ({fax_count/total_count*100:.1f}%)")
            print(f"📧 이메일 보유: {email_count}개 ({email_count/total_count*100:.1f}%)")
            print(f"🌐 URL 보유: {url_count}개 ({url_count/total_count*100:.1f}%)")
            print("=" * 50)
            
        except Exception as e:
            print(f"❌ 통계 계산 중 오류: {e}")

def main():
    """메인 실행 함수"""
    print("=" * 60)
    print("📊 JSON to Excel 변환기 (고급 크롤링 결과용)")
    print("=" * 60)
    
    # ContactDataExtractor 인스턴스 생성
    extractor = ContactDataExtractor()
    
    # 최신 JSON 파일 자동 찾기 (하드코딩 제거)
    try:
        json_file = find_latest_json_file()
        print(f"🔍 최신 JSON 파일 발견: {json_file}")
    except FileNotFoundError as e:
        print(f"❌ JSON 파일을 찾을 수 없습니다: {e}")
        return
    except Exception as e:
        print(f"❌ 파일 검색 중 오류: {e}")
        return
    
    # 타임스탬프 포함한 엑셀 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = f"church_data_enhanced_{timestamp}.xlsx"
    
    print(f"📂 입력 파일: {json_file}")
    print(f"💾 출력 파일: {excel_file}")
    
    # 파일 존재 확인
    if not os.path.exists(json_file):
        print(f"❌ 입력 파일을 찾을 수 없습니다: {json_file}")
        return
    
    # JSON to Excel 변환 실행
    success = asyncio.run(extractor.process_json_to_excel(json_file, excel_file, use_ai_validation=False))
    
    if success:
        print(f"\n🎉 변환 완료!")
        print(f"📁 생성된 파일: {excel_file}")
        
        # 결과 미리보기
        extractor.preview_excel_data(excel_file, num_rows=3)
        
        # 통계 출력
        extractor.count_excel_statistics(excel_file)
        
        # 파일 크기 정보
        try:
            file_size = os.path.getsize(excel_file) / (1024 * 1024)  # MB 단위
            print(f"📏 파일 크기: {file_size:.2f} MB")
        except:
            pass
    else:
        print("❌ 변환 실패")

if __name__ == "__main__":
    main()  # asyncio.run()은 이미 process_json_to_excel() 호출 시 사용됨