#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
크롤링 데이터 통계 분석기
JSON/Excel 파일의 크롤링 결과를 분석하고 다양한 통계를 제공하는 모듈
"""

import json
import os
import glob
import re
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from collections import Counter, defaultdict
import traceback

from settings import (
    KOREAN_AREA_CODES,
    AREA_CODE_LENGTH_RULES,
    PORTAL_EMAIL_DOMAINS,
    GOVERNMENT_EMAIL_SUFFIXES,
    EDUCATION_EMAIL_SUFFIXES,
    BUSINESS_EMAIL_SUFFIXES,
    RELIGIOUS_EMAIL_KEYWORDS,
    get_area_name,
    format_phone_number,
    extract_phone_area_code
)

# Excel 관련 라이브러리 (선택적)
try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl 없음 - Excel 분석 기능 비활성화")

class DataStatisticsAnalyzer:
    """크롤링 데이터 통계 분석 클래스"""
    
    def __init__(self):
        """초기화"""
        # constants.py에서 임포트한 상수들 사용
        # self.korean_area_codes는 이제 constants.KOREAN_AREA_CODES를 사용
        
        self.email_domains = {}
        self.phone_patterns = {}
        self.fax_patterns = {}
        
    def find_latest_files(self) -> Dict[str, Optional[str]]:
        """최신 JSON 및 Excel 파일 찾기"""
        files = {
            'json': None,
            'excel': None
        }
        
        # JSON 파일 찾기
        json_pattern = "raw_data_with_homepages_*.json"
        json_files = glob.glob(json_pattern)
        if json_files:
            files['json'] = max(json_files, key=os.path.getctime)
            print(f"📂 JSON 파일 발견: {files['json']}")
        
        # Excel 파일 찾기
        excel_pattern = "contact_data_*.xlsx"
        excel_files = glob.glob(excel_pattern)
        if excel_files:
            files['excel'] = max(excel_files, key=os.path.getctime)
            print(f"📊 Excel 파일 발견: {files['excel']}")
        
        return files
    
    def extract_phone_area_code(self, phone: str) -> Optional[str]:
        """전화번호에서 지역코드 추출"""
        if not phone:
            return None
        
        # 숫자만 추출
        digits = re.sub(r'[^\d]', '', phone)
        
        # 지역코드 매칭
        for code, area in self.korean_area_codes.items():
            if digits.startswith(code):
                return code
        
        return None
    
    def validate_korean_phone(self, phone: str) -> Dict[str, Any]:
        """한국 전화번호 유효성 검증"""
        if not phone:
            return {"is_valid": False, "reason": "번호 없음"}
        
        digits = re.sub(r'[^\d]', '', phone)
        
        # 길이 체크
        if len(digits) < 9 or len(digits) > 11:
            return {"is_valid": False, "reason": "길이 오류"}
        
        # 지역코드 체크
        area_code = self.extract_phone_area_code(phone)
        if not area_code:
            return {"is_valid": False, "reason": "지역코드 오류"}
        
        # 길이별 유효성 체크
        area_info = self.korean_area_codes[area_code]
        expected_lengths = {
            '02': [9, 10],  # 서울
            '070': [11],    # 인터넷전화
            '010': [11], '017': [11]  # 핸드폰
        }
        
        if area_code in expected_lengths:
            if len(digits) not in expected_lengths[area_code]:
                return {"is_valid": False, "reason": "길이 불일치"}
        else:
            # 기타 지역번호는 10-11자리
            if len(digits) not in [10, 11]:
                return {"is_valid": False, "reason": "길이 불일치"}
        
        return {
            "is_valid": True,
            "area_code": area_code,
            "area_name": area_info,
            "formatted": self.format_phone_number(digits, area_code)
        }
    
    def format_phone_number(self, digits: str, area_code: str) -> str:
        """전화번호 포맷팅 (constants의 함수 사용)"""
        return format_phone_number(digits, area_code)
    
    def extract_email_domain(self, email: str) -> Optional[str]:
        """이메일에서 도메인 추출"""
        if not email or '@' not in email:
            return None
        
        try:
            domain = email.split('@')[1].lower()
            return domain
        except:
            return None
    
    def categorize_email_domain(self, domain: str) -> str:
        """이메일 도메인 카테고리 분류 (constants의 상수들 사용)"""
        if not domain:
            return "기타"
        
        # constants의 PORTAL_EMAIL_DOMAINS 사용
        if domain in PORTAL_EMAIL_DOMAINS:
            return "포털"
        
        # constants의 GOVERNMENT_EMAIL_SUFFIXES 사용
        if any(domain.endswith(suffix) for suffix in GOVERNMENT_EMAIL_SUFFIXES):
            return "정부/공공"
        
        # constants의 EDUCATION_EMAIL_SUFFIXES 사용
        if any(domain.endswith(suffix) for suffix in EDUCATION_EMAIL_SUFFIXES):
            return "교육기관"
        
        # constants의 BUSINESS_EMAIL_SUFFIXES 사용
        if any(domain.endswith(suffix) for suffix in BUSINESS_EMAIL_SUFFIXES):
            return "기업/조직"
        
        # constants의 RELIGIOUS_EMAIL_KEYWORDS 사용
        if any(keyword in domain for keyword in RELIGIOUS_EMAIL_KEYWORDS):
            return "종교기관"
        
        return "기타"
    
    def analyze_json_data(self, json_file: str) -> Dict[str, Any]:
        """JSON 파일 데이터 분석"""
        print(f"🔍 JSON 데이터 분석 시작: {json_file}")
        
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 통계 초기화
            stats = {
                "file_info": {
                    "filename": json_file,
                    "analysis_time": datetime.now().isoformat()
                },
                "basic_stats": {},
                "categories": {},
                "contact_coverage": {},
                "quality_metrics": {},
                "geographic_distribution": {},
                "email_analysis": {}
            }
            
            # 카운터 초기화
            total_orgs = 0
            phone_count = 0
            fax_count = 0
            email_count = 0
            url_count = 0
            complete_contact_count = 0  # 전화+팩스+이메일 모두 있는 기관
            
            valid_phones = 0
            valid_faxes = 0
            
            phone_areas = Counter()
            fax_areas = Counter()
            email_domains = Counter()
            
            # 카테고리별 분석
            for category, organizations in data.items():
                category_stats = {
                    "count": len(organizations),
                    "phone_count": 0,
                    "fax_count": 0,
                    "email_count": 0,
                    "url_count": 0,
                    "complete_contact_count": 0  # 전화+팩스+이메일 모두 있는 기관
                }
                
                for org in organizations:
                    total_orgs += 1
                    
                    # 기본 연락처 정보
                    phone = org.get("phone", "")
                    fax = org.get("fax", "")
                    url = org.get("homepage", "")
                    
                    # 홈페이지 파싱 결과
                    homepage_content = org.get("homepage_content", {})
                    parsed_contact = homepage_content.get("parsed_contact", {})
                    
                    # 전화번호 분석
                    final_phone = phone
                    if not final_phone and parsed_contact.get("phones"):
                        final_phone = ", ".join(parsed_contact["phones"])
                    
                    if final_phone:
                        phone_count += 1
                        category_stats["phone_count"] += 1
                        
                        # 첫 번째 전화번호만 분석 (여러 개인 경우)
                        first_phone = final_phone.split(',')[0].strip()
                        phone_validation = self.validate_korean_phone(first_phone)
                        if phone_validation["is_valid"]:
                            valid_phones += 1
                            area_code = phone_validation["area_code"]
                            phone_areas[area_code] += 1
                    
                    # 팩스번호 분석
                    final_fax = fax
                    if not final_fax and parsed_contact.get("faxes"):
                        final_fax = ", ".join(parsed_contact["faxes"])
                    
                    if final_fax:
                        fax_count += 1
                        category_stats["fax_count"] += 1
                        
                        # 첫 번째 팩스번호만 분석 (여러 개인 경우)
                        first_fax = final_fax.split(',')[0].strip()
                        fax_validation = self.validate_korean_phone(first_fax)  # 팩스도 전화번호 검증 로직 사용
                        if fax_validation["is_valid"]:
                            valid_faxes += 1
                            area_code = fax_validation["area_code"]
                            fax_areas[area_code] += 1
                    
                    # 이메일 분석
                    final_email = ""
                    if parsed_contact.get("emails"):
                        final_email = ", ".join(parsed_contact["emails"])
                    
                    if final_email:
                        email_count += 1
                        category_stats["email_count"] += 1
                        
                        # 첫 번째 이메일만 분석 (여러 개인 경우)
                        first_email = final_email.split(',')[0].strip()
                        domain = self.extract_email_domain(first_email)
                        if domain:
                            email_domains[domain] += 1
                    
                    # URL 분석
                    if url:
                        url_count += 1
                        category_stats["url_count"] += 1
                    
                    # 완전한 연락처 정보 체크
                    if final_phone and final_fax and final_email:
                        complete_contact_count += 1
                        category_stats["complete_contact_count"] += 1
                
                stats["categories"][category] = category_stats
            
            # 기본 통계
            stats["basic_stats"] = {
                "total_organizations": total_orgs,
                "total_categories": len(data),
                "complete_contacts": complete_contact_count
            }
            
            # 연락처 커버리지
            stats["contact_coverage"] = {
                "phone": {"count": phone_count, "rate": round((phone_count / total_orgs * 100), 1) if total_orgs > 0 else 0},
                "fax": {"count": fax_count, "rate": round((fax_count / total_orgs * 100), 1) if total_orgs > 0 else 0},
                "email": {"count": email_count, "rate": round((email_count / total_orgs * 100), 1) if total_orgs > 0 else 0},
                "url": {"count": url_count, "rate": round((url_count / total_orgs * 100), 1) if total_orgs > 0 else 0}
            }
            
            # 품질 지표
            stats["quality_metrics"] = {
                "completeness_score": (complete_contact_count / total_orgs * 100) if total_orgs > 0 else 0,
                "phone_coverage": (phone_count / total_orgs * 100) if total_orgs > 0 else 0,
                "fax_coverage": (fax_count / total_orgs * 100) if total_orgs > 0 else 0,
                "email_coverage": (email_count / total_orgs * 100) if total_orgs > 0 else 0,
                "url_coverage": (url_count / total_orgs * 100) if total_orgs > 0 else 0,
                "phone_validity": (valid_phones / phone_count * 100) if phone_count > 0 else 0,
                "fax_validity": (valid_faxes / fax_count * 100) if fax_count > 0 else 0
            }
            
            # 지역 분포 (constants의 KOREAN_AREA_CODES 사용)
            all_areas = Counter()
            all_areas.update(phone_areas)
            all_areas.update(fax_areas)
            
            stats["geographic_distribution"] = {
                area_code: {
                    "area_name": KOREAN_AREA_CODES[area_code],  # constants 사용
                    "total_contacts": count,
                    "phone_count": phone_areas.get(area_code, 0),
                    "fax_count": fax_areas.get(area_code, 0)
                }
                for area_code, count in all_areas.most_common()
            }
            
            # 이메일 도메인 분석
            email_categories = Counter()
            for domain, count in email_domains.most_common(50):  # 상위 50개 도메인만
                category = self.categorize_email_domain(domain)
                email_categories[category] += count
            
            stats["email_analysis"] = {
                "top_domains": dict(email_domains.most_common(20)),
                "domain_categories": dict(email_categories),
                "total_unique_domains": len(email_domains)
            }
            
            print(f"✅ JSON 데이터 분석 완료")
            print(f"📊 총 기관: {total_orgs:,}개")
            print(f"📞 전화번호: {phone_count:,}개 ({phone_count/total_orgs*100:.1f}%)")
            print(f"📠 팩스번호: {fax_count:,}개 ({fax_count/total_orgs*100:.1f}%)")
            print(f"✉️ 이메일: {email_count:,}개 ({email_count/total_orgs*100:.1f}%)")
            print(f"🌐 홈페이지: {url_count:,}개 ({url_count/total_orgs*100:.1f}%)")
            
            return stats
            
        except Exception as e:
            print(f"❌ JSON 분석 실패: {e}")
            traceback.print_exc()
            return {}
    
    def analyze_excel_data(self, excel_file: str) -> Dict[str, Any]:
        """Excel 파일 데이터 분석"""
        if not EXCEL_AVAILABLE:
            print("⚠️ Excel 분석 불가능 - openpyxl 라이브러리 필요")
            return {}
        
        print(f"📊 Excel 데이터 분석 시작: {excel_file}")
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            stats = {
                "file_info": {
                    "filename": excel_file,
                    "file_size": os.path.getsize(excel_file),
                    "analysis_time": datetime.now().isoformat(),
                    "total_rows": ws.max_row - 1  # 헤더 제외
                },
                "column_analysis": {},
                "data_quality": {},
                "summary": {}
            }
            
            # 헤더 분석
            headers = [cell.value for cell in ws[1]]
            
            # 컬럼별 데이터 분석
            column_stats = {}
            for col_idx, header in enumerate(headers, 1):
                if not header:
                    continue
                
                values = []
                non_empty_count = 0
                
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value and str(cell_value).strip():
                        values.append(str(cell_value).strip())
                        non_empty_count += 1
                
                column_stats[header] = {
                    "total_values": len(values),
                    "non_empty_count": non_empty_count,
                    "fill_rate": (non_empty_count / (ws.max_row - 1) * 100) if ws.max_row > 1 else 0,
                    "unique_values": len(set(values)),
                    "sample_values": values[:5] if values else []
                }
                
                # 전화번호/팩스번호 컬럼 특별 분석
                if "전화" in header or "phone" in header.lower():
                    valid_count = 0
                    area_distribution = Counter()
                    
                    for value in values:
                        validation = self.validate_korean_phone(value)
                        if validation["is_valid"]:
                            valid_count += 1
                            area_distribution[validation["area_code"]] += 1
                    
                    column_stats[header]["validation"] = {
                        "valid_count": valid_count,
                        "validity_rate": (valid_count / len(values) * 100) if values else 0,
                        "area_distribution": dict(area_distribution.most_common(5))
                    }
                
                # 이메일 컬럼 특별 분석
                elif "이메일" in header or "email" in header.lower():
                    domain_distribution = Counter()
                    category_distribution = Counter()
                    
                    for value in values:
                        if '@' in value:
                            domain = self.extract_email_domain(value)
                            if domain:
                                domain_distribution[domain] += 1
                                category = self.categorize_email_domain(domain)
                                category_distribution[category] += 1
                    
                    column_stats[header]["analysis"] = {
                        "domain_distribution": dict(domain_distribution.most_common(5)),
                        "category_distribution": dict(category_distribution)
                    }
            
            stats["column_analysis"] = column_stats
            
            # 데이터 품질 분석
            total_rows = ws.max_row - 1
            complete_rows = 0
            
            for row_idx in range(2, ws.max_row + 1):
                row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, len(headers) + 1)]
                if all(data and str(data).strip() for data in row_data[:5]):  # 기본 5개 컬럼 체크
                    complete_rows += 1
            
            stats["data_quality"] = {
                "total_rows": total_rows,
                "complete_rows": complete_rows,
                "completeness_rate": (complete_rows / total_rows * 100) if total_rows > 0 else 0,
                "missing_data_rate": ((total_rows - complete_rows) / total_rows * 100) if total_rows > 0 else 0
            }
            
            # 요약
            stats["summary"] = {
                "total_records": total_rows,
                "total_columns": len(headers),
                "data_completeness": round((complete_rows / total_rows * 100), 1) if total_rows > 0 else 0,
                "column_fill_rates": {
                    header: round(data["fill_rate"], 1) 
                    for header, data in column_stats.items()
                }
            }
            
            print(f"✅ Excel 분석 완료: {total_rows}개 레코드 분석")
            return stats
            
        except Exception as e:
            print(f"❌ Excel 분석 중 오류: {e}")
            print(f"🔍 상세 오류: {traceback.format_exc()}")
            return {}
    
    def generate_comprehensive_report(self, json_stats: Dict, excel_stats: Dict) -> Dict[str, Any]:
        """종합 분석 리포트 생성"""
        print("📋 종합 분석 리포트 생성 중...")
        
        report = {
            "report_info": {
                "generated_at": datetime.now().isoformat(),
                "analysis_type": "comprehensive",
                "data_sources": []
            },
            "executive_summary": {},
            "detailed_analysis": {},
            "recommendations": [],
            "data_quality_score": 0
        }
        
        # 데이터 소스 정보
        if json_stats:
            report["report_info"]["data_sources"].append({
                "type": "JSON",
                "file": json_stats.get("file_info", {}).get("filename", ""),
                "records": json_stats.get("summary", {}).get("total_organizations", 0)
            })
        
        if excel_stats:
            report["report_info"]["data_sources"].append({
                "type": "Excel", 
                "file": excel_stats.get("file_info", {}).get("filename", ""),
                "records": excel_stats.get("summary", {}).get("total_records", 0)
            })
        
        # 경영진 요약
        if json_stats:
            summary = json_stats.get("summary", {})
            quality = json_stats.get("quality_metrics", {})
            
            report["executive_summary"] = {
                "total_organizations": summary.get("total_organizations", 0),
                "contact_coverage": {
                    "phone": summary.get("coverage_rates", {}).get("phone", 0),
                    "fax": summary.get("coverage_rates", {}).get("fax", 0),
                    "email": summary.get("coverage_rates", {}).get("email", 0),
                    "url": summary.get("coverage_rates", {}).get("url", 0)
                },
                "data_quality": {
                    "phone_validity": quality.get("phone_validity", 0),
                    "fax_validity": quality.get("fax_validity", 0),
                    "completeness": quality.get("completeness_score", 0)
                },
                "top_regions": list(json_stats.get("geographic_distribution", {}).keys())[:5]
            }
        
        # 상세 분석
        report["detailed_analysis"] = {
            "json_analysis": json_stats,
            "excel_analysis": excel_stats
        }
        
        # 권장사항 생성
        recommendations = []
        
        if json_stats:
            quality = json_stats.get("quality_metrics", {})
            
            if quality.get("phone_validity", 0) < 90:
                recommendations.append({
                    "priority": "high",
                    "category": "data_quality",
                    "issue": "전화번호 유효성 낮음",
                    "description": f"전화번호 유효성이 {quality.get('phone_validity', 0):.1f}%로 낮습니다.",
                    "action": "전화번호 형식 검증 로직 강화 필요"
                })
            
            if quality.get("completeness_score", 0) < 50:
                recommendations.append({
                    "priority": "medium",
                    "category": "data_coverage",
                    "issue": "연락처 정보 완성도 낮음",
                    "description": f"완전한 연락처 정보를 가진 기관이 {quality.get('completeness_score', 0):.1f}%에 불과합니다.",
                    "action": "추가 크롤링 소스 발굴 또는 수동 보완 필요"
                })
            
            if quality.get("email_coverage", 0) < 30:
                recommendations.append({
                    "priority": "medium",
                    "category": "data_coverage", 
                    "issue": "이메일 정보 부족",
                    "description": f"이메일 정보 보유율이 {quality.get('email_coverage', 0):.1f}%로 낮습니다.",
                    "action": "이메일 추출 알고리즘 개선 또는 추가 소스 활용"
                })
        
        report["recommendations"] = recommendations
        
        # 데이터 품질 점수 계산
        if json_stats:
            quality = json_stats.get("quality_metrics", {})
            coverage_score = (
                quality.get("phone_coverage", 0) + 
                quality.get("fax_coverage", 0) + 
                quality.get("email_coverage", 0) + 
                quality.get("url_coverage", 0)
            ) / 4
            
            validity_score = (
                quality.get("phone_validity", 0) + 
                quality.get("fax_validity", 0)
            ) / 2
            
            completeness_score = quality.get("completeness_score", 0)
            
            overall_score = (coverage_score * 0.4 + validity_score * 0.3 + completeness_score * 0.3)
            report["data_quality_score"] = round(overall_score, 1)
        
        print("✅ 종합 리포트 생성 완료")
        return report
    
    def print_summary_report(self, stats: Dict[str, Any]):
        """요약 리포트 콘솔 출력"""
        print("\n" + "=" * 80)
        print("📊 크롤링 데이터 통계 분석 리포트")
        print("=" * 80)
        
        if "summary" in stats:
            summary = stats["summary"]
            print(f"\n📈 기본 통계:")
            print(f"   총 기관 수: {summary.get('total_organizations', 0):,}개")
            print(f"   카테고리 수: {summary.get('total_categories', 0)}개")
            
            print(f"\n📞 연락처 보유 현황:")
            coverage = summary.get("coverage_rates", {})
            counts = summary.get("contact_counts", {})
            print(f"   전화번호: {counts.get('phone', 0):,}개 ({coverage.get('phone', 0)}%)")
            print(f"   팩스번호: {counts.get('fax', 0):,}개 ({coverage.get('fax', 0)}%)")
            print(f"   이메일: {counts.get('email', 0):,}개 ({coverage.get('email', 0)}%)")
            print(f"   홈페이지: {counts.get('url', 0):,}개 ({coverage.get('url', 0)}%)")
            
            print(f"\n🎯 데이터 품질 지표:")
            quality = summary.get("quality_scores", {})
            print(f"   전화번호 유효성: {quality.get('phone_validity', 0)}%")
            print(f"   팩스번호 유효성: {quality.get('fax_validity', 0)}%")
            print(f"   정보 완성도: {quality.get('completeness', 0)}%")
        
        if "geographic_distribution" in stats:
            print(f"\n🗺️ 지역별 분포 (상위 5개):")
            geo_dist = stats["geographic_distribution"]
            for i, (code, info) in enumerate(list(geo_dist.items())[:5]):
                area_name = info.get("area_name", "알 수 없음")
                total = info.get("total_contacts", 0)
                print(f"   {i+1}. {area_name}({code}): {total}개 연락처")
        
        if "contact_analysis" in stats:
            email_stats = stats["contact_analysis"].get("email_stats", {})
            if "category_distribution" in email_stats:
                print(f"\n📧 이메일 도메인 분류:")
                for category, count in email_stats["category_distribution"].items():
                    print(f"   {category}: {count}개")
        
        print("\n" + "=" * 80)
    
    def save_report_to_json(self, report: Dict[str, Any], filename: Optional[str] = None) -> str:
        """리포트를 JSON 파일로 저장"""
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"data_analysis_report_{timestamp}.json"
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(report, f, ensure_ascii=False, indent=2)
            
            print(f"💾 분석 리포트 저장: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ 리포트 저장 실패: {e}")
            return ""
    
    def get_api_summary(self, stats: Dict[str, Any]) -> Dict[str, Any]:
        """웹 API용 요약 데이터 생성"""
        if not stats or "summary" not in stats:
            return {}
        
        summary = stats["summary"]
        quality = stats.get("quality_metrics", {})
        
        api_data = {
            "basic_stats": {
                "total_organizations": summary.get("total_organizations", 0),
                "total_categories": summary.get("total_categories", 0)
            },
            "contact_coverage": {
                "phone": {
                    "count": summary.get("contact_counts", {}).get("phone", 0),
                    "rate": summary.get("coverage_rates", {}).get("phone", 0)
                },
                "fax": {
                    "count": summary.get("contact_counts", {}).get("fax", 0),
                    "rate": summary.get("coverage_rates", {}).get("fax", 0)
                },
                "email": {
                    "count": summary.get("contact_counts", {}).get("email", 0),
                    "rate": summary.get("coverage_rates", {}).get("email", 0)
                },
                "url": {
                    "count": summary.get("contact_counts", {}).get("url", 0),
                    "rate": summary.get("coverage_rates", {}).get("url", 0)
                }
            },
            "quality_metrics": {
                "phone_validity": quality.get("phone_validity", 0),
                "fax_validity": quality.get("fax_validity", 0),
                "completeness_score": quality.get("completeness_score", 0),
                "overall_score": stats.get("data_quality_score", 0)
            },
            "top_regions": [],
            "email_categories": {}
        }
        
        # 상위 지역 정보
        if "geographic_distribution" in stats:
            geo_dist = stats["geographic_distribution"]
            api_data["top_regions"] = [
                {
                    "code": code,
                    "name": info.get("area_name", ""),
                    "contact_count": info.get("total_contacts", 0)
                }
                for code, info in list(geo_dist.items())[:5]
            ]
        
        # 이메일 카테고리 분포
        if "contact_analysis" in stats:
            email_stats = stats["contact_analysis"].get("email_stats", {})
            if "category_distribution" in email_stats:
                api_data["email_categories"] = email_stats["category_distribution"]
        
        return api_data

def main():
    """메인 실행 함수"""
    print("=" * 80)
    print("📊 크롤링 데이터 통계 분석기")
    print("=" * 80)
    
    analyzer = DataStatisticsAnalyzer()
    
    # 최신 파일 찾기
    files = analyzer.find_latest_files()
    
    if not files['json'] and not files['excel']:
        print("❌ 분석할 파일을 찾을 수 없습니다.")
        print("💡 raw_data_with_homepages_*.json 또는 contact_data_*.xlsx 파일이 필요합니다.")
        return
    
    # 분석 실행
    json_stats = {}
    excel_stats = {}
    
    if files['json']:
        json_stats = analyzer.analyze_json_data(files['json'])
    
    if files['excel']:
        excel_stats = analyzer.analyze_excel_data(files['excel'])
    
    # 종합 리포트 생성
    if json_stats or excel_stats:
        comprehensive_report = analyzer.generate_comprehensive_report(json_stats, excel_stats)
        
        # 콘솔 출력
        if json_stats:
            analyzer.print_summary_report(json_stats)
        
        # JSON 리포트 저장
        report_file = analyzer.save_report_to_json(comprehensive_report)
        
        # API용 요약 데이터 생성 (예시)
        if json_stats:
            api_summary = analyzer.get_api_summary(json_stats)
            print(f"\n🌐 API용 요약 데이터 생성 완료 (웹 연동 준비)")
            print(f"   기관 수: {api_summary.get('basic_stats', {}).get('total_organizations', 0)}")
            print(f"   품질 점수: {api_summary.get('quality_metrics', {}).get('overall_score', 0)}")
        
        print(f"\n✅ 분석 완료!")
        print(f"📁 상세 리포트: {report_file}")
    else:
        print("❌ 분석할 데이터가 없습니다.")

if __name__ == "__main__":
    main() 